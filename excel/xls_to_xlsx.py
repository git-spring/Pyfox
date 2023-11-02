
import xlrd
import openpyxl


def xls_to_xlsx(src_file_path):
    """
    将指定的xls文件转换为xlsx格式
    :param src_file_path: xls文件路径
    :return 新的文件路径名称
    """
    import os
    folder_path = os.path.dirname(src_file_path)
    file_name = os.path.basename(src_file_path)
    name,suffix = file_name.split('.')
    new_excel_file_path = f'{folder_path}/{name}.xlsx'  # 新文件的路径名称

    if suffix == 'xlsx':                                 # 如果是 xlsx 格式的,不用处理
        return src_file_path

    wb_xls = xlrd.open_workbook(f'{src_file_path}')      # xlrd 打开 xls 文件
    wb_xlsx = openpyxl.Workbook()                        # 创建 xlsx 文件,用来保存新的文件
    sheet_names = wb_xls.sheet_names()                   # 获取所有sheet名称

    for sheet in sheet_names :                           # 循环读取sheet表格的内容,并写入新的文件
        ws_xls = wb_xls.sheet_by_name(sheet)
        nrows = ws_xls.nrows                             # sheet工作表的行
        sheet_msg_xls = []
        for i in range(nrows):                           # 读取xls的sheet内容存到写入列表中
            sheet_msg_xls.append(ws_xls.row_values(i))

        ws_xlxs = wb_xlsx.create_sheet(f'{sheet}')       # 在xlsx创建同名sheet
        for row in sheet_msg_xls:                        # 把xls sheet内容写入到 xlsx同名sheet中
            ws_xlxs.append(row)

        wb_xlsx.save(f'{new_excel_file_path}')           # 保存xlsx格式文件

    wb = openpyxl.load_workbook(new_excel_file_path)     # 打开新的 xlsx格式文件
    del wb[wb.sheetnames[0]]                             # 删除第一个默认创建的Sheet页
    wb.save(new_excel_file_path)
    return new_excel_file_path



xls_to_xlsx('D:/test/中指院_房地产数据_销售榜1111-201801-202309.xls')


